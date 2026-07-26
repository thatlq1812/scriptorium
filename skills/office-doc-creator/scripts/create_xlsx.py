#!/usr/bin/env python3
"""Create a .xlsx from a simple JSON content spec.

Usage:
    python create_xlsx.py <content.json> <output.xlsx>

content.json shape:
{
  "sheets": [
    {"name": "Sheet1", "headers": ["A", "B"], "rows": [[1, 2], [3, 4]]}
  ]
}
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook


def build(content: dict, output_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_spec in content.get("sheets", []):
        ws = wb.create_sheet(title=sheet_spec.get("name", "Sheet"))
        headers = sheet_spec.get("headers", [])
        if headers:
            ws.append(headers)
        for row in sheet_spec.get("rows", []):
            ws.append(row)

    if not wb.sheetnames:
        wb.create_sheet(title="Sheet1")

    wb.save(str(output_path))


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("content_json", type=Path)
    parser.add_argument("output_xlsx", type=Path)
    args = parser.parse_args()

    content = json.loads(args.content_json.read_text(encoding="utf-8"))
    build(content, args.output_xlsx)
    print(f"OK: wrote {args.output_xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
